from flask import Flask, request, render_template
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

@app.route('/')
def form():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']

    file_name = "data.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email"])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([name, email])
    wb.save(file_name)

    return render_template("success.html", name=name)

if __name__ == "__main__":
    app.run(debug=True)